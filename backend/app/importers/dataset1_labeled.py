from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook

from app.importers.base import BaseImporter, ImporterError

JUNIOR_COMPETENCY_HEADERS = [
    "抽象能力",
    "运算能力",
    "几何直观",
    "空间观念",
    "推理能力",
    "数据观念",
    "模型观念",
    "应用意识",
    "创新意识",
]
SENIOR_COMPETENCY_HEADERS = [
    "数学抽象",
    "逻辑推理",
    "数学建模",
    "直观想象",
    "数学运算",
    "数据分析",
]
COGNITIVE_HEADERS = ["识记", "理解", "应用", "分析", "综合", "评价"]
QUESTION_ID_HEADERS = {"初中题目id", "高中题目id", "题目id"}
QUESTION_TEXT_HEADERS = {"试题内容"}
KNOWLEDGE_HEADERS = {"知识点"}
ALL_COMPETENCY_HEADERS = JUNIOR_COMPETENCY_HEADERS + SENIOR_COMPETENCY_HEADERS
GRADE_INDEX_BY_PREFIX = {
    "1": 7,
    "2": 8,
    "3": 9,
    "4": 10,
    "5": 11,
    "6": 12,
}
HEADER_ALIASES = {
    "1（识记）": "识记",
    "2（理解）": "理解",
    "3（应用）": "应用",
    "4（分析）": "分析",
    "5（综合）": "综合",
    "6（评价）": "评价",
}


class Dataset1LabeledImporter(BaseImporter):
    record_type = "question"

    def parse(self, file_path: Path) -> list[dict]:
        if not file_path.exists():
            raise ImporterError(f"File not found: {file_path}")

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 3:
            raise ImporterError("Dataset1 workbook does not contain expected header rows.")

        data_start_index = self._find_data_start_index(rows)
        if data_start_index is None:
            raise ImporterError("Dataset1 workbook does not contain any question rows.")

        merged_headers = self._merge_headers(rows[:data_start_index])

        records: list[dict] = []
        for row in rows[data_start_index:]:
            if not row or row[0] is None:
                continue

            payload = dict(zip(merged_headers, row, strict=False))
            question_id = self._first_payload_value(payload, QUESTION_ID_HEADERS)
            if question_id is None:
                continue

            question_id_text = str(question_id).strip()
            grade_index = self._infer_grade_index(question_id_text)
            edu_stage = self._infer_edu_stage(grade_index)
            competency_headers = (
                JUNIOR_COMPETENCY_HEADERS if edu_stage == "junior" else SENIOR_COMPETENCY_HEADERS
            )

            competencies = {
                header: int(payload.get(header) or 0)
                for header in competency_headers
            }
            cognitive_levels = {
                header: int(payload.get(header) or 0)
                for header in COGNITIVE_HEADERS
            }
            records.append(
                {
                    "source_record_key": question_id_text,
                    "raw_payload": {
                        "question_id": question_id_text,
                        "question_text": self._first_payload_value(payload, QUESTION_TEXT_HEADERS),
                        "knowledge_text": self._first_payload_value(payload, KNOWLEDGE_HEADERS),
                        "gradeIndex": grade_index,
                        "edu_stage": edu_stage,
                        "cognitive_levels": cognitive_levels,
                        "competencies": competencies,
                    },
                }
            )

        return records

    def _find_data_start_index(self, rows: list[tuple]) -> int | None:
        for index, row in enumerate(rows):
            if not row or row[0] is None:
                continue
            cell_text = str(row[0]).strip()
            if re.match(r"^Q_\d", cell_text):
                return index
        return None

    def _merge_headers(self, header_rows: list[tuple]) -> list[str | None]:
        width = max(len(row) for row in header_rows)
        merged_headers: list[str | None] = []
        known_leaf_headers = set(
            QUESTION_ID_HEADERS
            | QUESTION_TEXT_HEADERS
            | KNOWLEDGE_HEADERS
            | set(COGNITIVE_HEADERS)
            | set(ALL_COMPETENCY_HEADERS)
        )

        for column_index in range(width):
            parts: list[str] = []
            for row in header_rows:
                if column_index >= len(row):
                    continue
                value = row[column_index]
                if value is None:
                    continue
                normalized = self._normalize_header(str(value))
                if normalized:
                    parts.append(normalized)

            chosen = None
            for part in reversed(parts):
                if part in known_leaf_headers:
                    chosen = part
                    break
            if chosen is None and parts:
                chosen = parts[-1]
            merged_headers.append(chosen)
        return merged_headers

    def _normalize_header(self, value: str) -> str:
        compact = "".join(str(value).split())
        compact = compact.replace("\u3000", "")
        compact = HEADER_ALIASES.get(compact, compact)
        return compact

    def _first_payload_value(self, payload: dict, headers: set[str]):
        for key, value in payload.items():
            if key in headers:
                return value
        return None

    def _infer_grade_index(self, question_id: str) -> int | None:
        match = re.match(r"^Q_(\d)", question_id)
        if match is None:
            return None
        return GRADE_INDEX_BY_PREFIX.get(match.group(1))

    def _infer_edu_stage(self, grade_index: int | None) -> str | None:
        if grade_index is None:
            return None
        if 7 <= grade_index <= 9:
            return "junior"
        if 10 <= grade_index <= 12:
            return "senior"
        return None
