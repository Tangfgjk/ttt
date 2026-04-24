from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.importers.base import BaseImporter, ImporterError


class Dataset3ExamSheetImporter(BaseImporter):
    record_type = "exam_response"

    def parse(self, file_path: Path) -> list[dict]:
        if not file_path.exists():
            raise ImporterError(f"File not found: {file_path}")

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        if not workbook.sheetnames:
            raise ImporterError("Dataset3 workbook does not contain any worksheet.")

        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(item) if item is not None else "" for item in next(rows)]
        except StopIteration as exc:
            raise ImporterError("Dataset3 workbook is empty.") from exc

        records: list[dict] = []
        for row in rows:
            if not row or row[21] is None or row[22] is None or row[7] is None:
                continue
            payload = dict(zip(headers, row, strict=False))
            record_key = f"{payload['考试ID']}::{payload['学生编号']}::{payload['题目ID']}"
            records.append(
                {
                    "source_record_key": record_key,
                    "raw_payload": payload,
                }
            )

        return records
