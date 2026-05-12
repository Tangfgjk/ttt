from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

from fastapi import UploadFile
from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

from app.db.base import Base
from app.models.assessment import Exam, QuestionGoldCompetency, QuestionGoldLabel, StudentQuestionResponse
from app.models.dictionary import CognitiveLevel, Competency
from app.models.dictionary import Grade, QuestionType, Subject
from app.models.imports import DataSource
from app.models.question import Question, QuestionExternalRef
from app.services.import_service import ImportService


def _build_session() -> Session:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, future=True)()


def _seed_import_context(db: Session) -> None:
    db.add_all(
        [
            Subject(code="math", name="数学"),
            Grade(grade_index=8, grade_code="grade_8", grade_name="八年级", edu_stage="junior"),
            Grade(grade_index=7, grade_code="grade_7", grade_name="七年级", edu_stage="junior"),
            Grade(grade_index=9, grade_code="grade_9", grade_name="九年级", edu_stage="junior"),
            Grade(grade_index=10, grade_code="grade_10", grade_name="高一", edu_stage="senior"),
            Grade(grade_index=11, grade_code="grade_11", grade_name="高二", edu_stage="senior"),
            Grade(grade_index=12, grade_code="grade_12", grade_name="高三", edu_stage="senior"),
            QuestionType(code="select_single", name="单选题", base_type_index=1),
            QuestionType(code="single_choice", name="单选题", base_type_index=1),
            QuestionType(code="fill_blank", name="填空题", base_type_index=2),
            DataSource(
                code="dataset2_question_json",
                name="Dataset2",
                source_type="json",
                description="test source",
            ),
            DataSource(
                code="dataset1_labeled",
                name="Dataset1",
                source_type="excel",
                description="gold label source",
            ),
            DataSource(
                code="dataset3_exam_sheet",
                name="Dataset3",
                source_type="excel",
                description="exam sheet source",
            ),
            CognitiveLevel(code="remember", name="识记", level_order=1),
            CognitiveLevel(code="understand", name="理解", level_order=2),
            CognitiveLevel(code="apply", name="应用", level_order=3),
            CognitiveLevel(code="analyze", name="分析", level_order=4),
            CognitiveLevel(code="synthesize", name="综合", level_order=5),
            CognitiveLevel(code="evaluate", name="评价", level_order=6),
            Competency(code="abstraction", name="抽象能力", display_order=1),
            Competency(code="operation", name="运算能力", display_order=2),
            Competency(code="geometric_intuition", name="几何直观", display_order=3),
            Competency(code="spatial_conception", name="空间观念", display_order=4),
            Competency(code="reasoning", name="推理能力", display_order=5),
            Competency(code="data_consciousness", name="数据观念", display_order=6),
            Competency(code="model_consciousness", name="模型观念", display_order=7),
            Competency(code="application_awareness", name="应用意识", display_order=8),
            Competency(code="innovation_awareness", name="创新意识", display_order=9),
            Competency(code="mathematical_abstraction", name="数学抽象", display_order=10),
            Competency(code="logical_reasoning", name="逻辑推理", display_order=11),
            Competency(code="mathematical_modeling", name="数学建模", display_order=12),
            Competency(code="intuitive_imagination", name="直观想象", display_order=13),
            Competency(code="mathematical_operation", name="数学运算", display_order=14),
            Competency(code="data_analysis", name="数据分析", display_order=15),
        ]
    )
    db.commit()


def _write_dataset2_json(path: Path, *, exercise_id: str, stem_text: str) -> None:
    payload = {
        "exerciseID": exercise_id,
        "baseTypeIndex": 1,
        "blankCount": 1,
        "difficulty": 1,
        "exerciseType": "select_single",
        "gradeIndex": 8,
        "question": stem_text,
        "queAns": "D",
        "solution": "因为 D 正确。",
        "subQueNum": 0,
        "subQues": [],
        "subjectCategory": "math",
        "tags": [],
        "queCtlgs": [],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _write_dataset1_xlsx(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "初中题目id",
            "试题内容",
            "知识点",
            "认知层级水平",
            None,
            None,
            None,
            None,
            None,
            "核心素养",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    sheet.append(
        [
            None,
            None,
            None,
            "识记",
            "理解",
            "应用",
            "分析",
            "综合",
            "评价",
            "抽象能力",
            "运算能力",
            "几何直观",
            "空间观念",
            "推理能力",
            "数据观念",
            "模型观念",
            "应用意识",
            "创新意识",
        ]
    )
    sheet.append(
        [
            "Q_1001",
            "$2022$ 的倒数是（ ） A. -2022 B. 2022 C. 1/2022 D. -1/2022",
            "倒数",
            1,
            None,
            None,
            None,
            None,
            None,
            None,
            1,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    workbook.save(path)


def _write_dataset1_senior_xlsx(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "高中题目id",
            "知识点",
            "试题内容",
            "认知层级水平",
            None,
            None,
            None,
            None,
            None,
            "核心素养",
            None,
            None,
            None,
            None,
            None,
        ]
    )
    sheet.append([None] * 15)
    sheet.append(
        [
            None,
            None,
            None,
            "1\n（识记）",
            "2\n（理解）",
            "3\n（应用）",
            "4\n（分析）",
            "5\n（综合）",
            "6\n（评价）",
            "数学抽象",
            "逻辑推理",
            "数学建模",
            "直观想象",
            "数学运算",
            "数据分析",
        ]
    )
    sheet.append(
        [
            "Q_4001",
            "集合的交集运算",
            "已知集合 A={x|-1<=x<=2}, B={x∈N|x<2}, 则 A∩B=( )",
            None,
            1,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            1,
            None,
        ]
    )
    sheet.append(
        [
            "Q_5001",
            "函数建模",
            "某函数模型描述销量与价格关系，请判断最优定价。",
            None,
            None,
            None,
            1,
            None,
            None,
            2,
            None,
            1,
            None,
            None,
            None,
        ]
    )
    workbook.save(path)


def _write_dataset3_xlsx(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "测试时间",
            "测试名称",
            "测试学科",
            "测试年级",
            "测试类型",
            "测试班级",
            "测试总分",
            "学生编号",
            "测试得分",
            "题号",
            "自定义题号",
            "题目内容",
            "题干（子题）",
            "题型标签",
            "难度标签",
            "知识点标签",
            "章节标签",
            "题目答案",
            "题目分值",
            "学生作答",
            "题目得分",
            "考试ID",
            "题目ID",
            "知识点ID",
            "章节标签ID",
            "子题答案",
            "年级编号",
            "班级编号",
            "班级序号",
            "考试编号",
            "学期",
            "考试编号及名称",
        ]
    )
    sheet.append(
        [
            "2024-10-09T00:00:00",
            "七年级A-STEM综合素养评价-数学",
            "math",
            7,
            "月考",
            "八(1)班",
            100,
            3149240297,
            62,
            1,
            "1",
            "<p>-2023的相反数是( )</p><p>A.2023</p><p>B.-1/2023</p><p>C.1/2023</p><p>D.-2023</p>",
            "",
            "单选题",
            1,
            "求相反数",
            "1 认识有理数",
            "A",
            3,
            "A",
            3,
            10,
            "ZLDz5f6CNxWnZD3eXrw",
            "5eb8f00beb16623401483fd9",
            "BhblD84mYFndI0Vrsae",
            "",
            31490324,
            3149032401,
            1,
            31492409230102,
            "上学期",
            "31492409230102-七年级A-STEM综合素养评价-数学",
        ]
    )
    workbook.save(path)


def test_dataset2_import_creates_new_question(tmp_path: Path) -> None:
    db = _build_session()
    _seed_import_context(db)
    file_path = tmp_path / "dataset2.json"
    _write_dataset2_json(file_path, exercise_id="Q_001", stem_text="已知一次函数 y=2x+3，当 x=1 时求函数值。")

    result = ImportService(db).import_local_file("dataset2_question_json", str(file_path))
    detail = ImportService(db).get_batch_detail(result.batch.id)

    assert detail.summary.created_new_question == 1
    assert detail.summary.failed == 0
    assert detail.records[0].parse_status == "CREATED_NEW_QUESTION"
    assert detail.records[0].normalized_question_id is not None
    assert db.scalar(select(Question).limit(1)) is not None


def test_dataset2_import_matches_by_content_hash_for_new_external_id(tmp_path: Path) -> None:
    db = _build_session()
    _seed_import_context(db)
    service = ImportService(db)

    first = tmp_path / "dataset2_first.json"
    second = tmp_path / "dataset2_second.json"
    stem_text = "下列命题是假命题的是（ ）A. 三角形内角和是180度 B. 多边形外角和是360度 C. 两直线平行内错角相等 D. 1+1=3"
    _write_dataset2_json(first, exercise_id="Q_101", stem_text=stem_text)
    _write_dataset2_json(second, exercise_id="Q_102", stem_text=stem_text)

    first_result = service.import_local_file("dataset2_question_json", str(first))
    second_result = service.import_local_file("dataset2_question_json", str(second))
    first_detail = service.get_batch_detail(first_result.batch.id)
    second_detail = service.get_batch_detail(second_result.batch.id)

    assert first_detail.records[0].parse_status == "CREATED_NEW_QUESTION"
    assert second_detail.records[0].parse_status == "MATCHED_BY_CONTENT_HASH"
    assert second_detail.records[0].normalized_question_id == first_detail.records[0].normalized_question_id

    refs = list(db.scalars(select(QuestionExternalRef).order_by(QuestionExternalRef.id.asc())))
    assert len(refs) == 2


def test_dataset2_folder_import_merges_multiple_json_files_into_one_batch() -> None:
    db = _build_session()
    _seed_import_context(db)
    service = ImportService(db)

    payload_one = {
        "exerciseID": "FOLDER_001",
        "baseTypeIndex": 1,
        "blankCount": 0,
        "difficulty": 1,
        "exerciseType": "select_single",
        "gradeIndex": 8,
        "question": "已知一次函数 y=2x+3，当 x=2 时函数值是多少？",
        "queAns": "7",
        "solution": "将 x=2 代入即可。",
        "subQueNum": 0,
        "subQues": [],
        "subjectCategory": "math",
        "tags": [],
        "queCtlgs": [],
    }
    payload_two = {
        "exerciseID": "FOLDER_002",
        "baseTypeIndex": 1,
        "blankCount": 0,
        "difficulty": 1,
        "exerciseType": "select_single",
        "gradeIndex": 8,
        "question": "已知一次函数 y=2x+3，当 x=2 时函数值是多少？",
        "queAns": "7",
        "solution": "将 x=2 代入即可。",
        "subQueNum": 0,
        "subQues": [],
        "subjectCategory": "math",
        "tags": [],
        "queCtlgs": [],
    }

    upload_one = UploadFile(
        filename="folder_a.json",
        file=BytesIO(json.dumps(payload_one, ensure_ascii=False).encode("utf-8")),
    )
    upload_two = UploadFile(
        filename="folder_b.json",
        file=BytesIO(json.dumps(payload_two, ensure_ascii=False).encode("utf-8")),
    )

    result = service.import_uploaded_files(
        "dataset2_question_json",
        [upload_one, upload_two],
        folder_name="题目合并",
    )
    detail = service.get_batch_detail(result.batch.id)

    assert detail.batch.file_name == "题目合并 (2 files)"
    assert detail.summary.total_records == 2
    assert detail.summary.created_new_question == 1
    assert detail.summary.matched_by_content_hash == 1
    assert detail.records[0].normalized_question_id == detail.records[1].normalized_question_id


def test_dataset2_folder_import_can_append_chunks_to_same_batch() -> None:
    db = _build_session()
    _seed_import_context(db)
    service = ImportService(db)

    batch = service.initialize_upload_batch(
        "dataset2_question_json",
        file_name="题目合并 (2 files)",
    )

    first_payload = {
        "exerciseID": "CHUNK_001",
        "baseTypeIndex": 1,
        "blankCount": 0,
        "difficulty": 1,
        "exerciseType": "select_single",
        "gradeIndex": 8,
        "question": "若 a+b=5，且 a=2，则 b 的值是多少？",
        "queAns": "3",
        "solution": "5-2=3。",
        "subQueNum": 0,
        "subQues": [],
        "subjectCategory": "math",
        "tags": [],
        "queCtlgs": [],
    }
    second_payload = {
        "exerciseID": "CHUNK_002",
        "baseTypeIndex": 1,
        "blankCount": 0,
        "difficulty": 1,
        "exerciseType": "select_single",
        "gradeIndex": 8,
        "question": "若 a+b=5，且 a=2，则 b 的值是多少？",
        "queAns": "3",
        "solution": "5-2=3。",
        "subQueNum": 0,
        "subQues": [],
        "subjectCategory": "math",
        "tags": [],
        "queCtlgs": [],
    }

    first_upload = UploadFile(
        filename="chunk_1.json",
        file=BytesIO(json.dumps(first_payload, ensure_ascii=False).encode("utf-8")),
    )
    second_upload = UploadFile(
        filename="chunk_2.json",
        file=BytesIO(json.dumps(second_payload, ensure_ascii=False).encode("utf-8")),
    )

    first_result = service.append_uploaded_files_to_batch(batch.id, [first_upload], finalize=False)
    first_detail = service.get_batch_detail(batch.id)
    second_result = service.append_uploaded_files_to_batch(batch.id, [second_upload], finalize=True)
    detail = service.get_batch_detail(batch.id)

    assert first_result.summary.total_records == 1
    assert first_detail.batch.import_status == "RUNNING"
    assert second_result.batch.import_status == "SUCCESS"
    assert detail.summary.total_records == 2
    assert detail.summary.created_new_question == 1
    assert detail.summary.matched_by_content_hash == 1


def test_dataset2_import_reuses_catalog_when_school_code_differs(tmp_path: Path) -> None:
    db = _build_session()
    _seed_import_context(db)
    service = ImportService(db)

    file_path = tmp_path / "dataset2_catalog_reuse.json"
    payload = [
        {
            "exerciseID": "CAT_001",
            "baseTypeIndex": 1,
            "blankCount": 0,
            "difficulty": 1,
            "exerciseType": "select_single",
            "gradeIndex": 8,
            "question": "已知 2+3 的结果是几？",
            "queAns": "5",
            "solution": "",
            "subQueNum": 0,
            "subQues": [],
            "subjectCategory": "math",
            "tags": [],
            "queCtlgs": [
                {
                    "catalogId": "catalog_same",
                    "queId": "CAT_001",
                    "schCode": "3019",
                    "textbookId": "textbook_same",
                }
            ],
        },
        {
            "exerciseID": "CAT_002",
            "baseTypeIndex": 1,
            "blankCount": 0,
            "difficulty": 1,
            "exerciseType": "select_single",
            "gradeIndex": 8,
            "question": "已知 4+1 的结果是几？",
            "queAns": "5",
            "solution": "",
            "subQueNum": 0,
            "subQues": [],
            "subjectCategory": "math",
            "tags": [],
            "queCtlgs": [
                {
                    "catalogId": "catalog_same",
                    "queId": "CAT_002",
                    "schCode": "3030",
                    "textbookId": "textbook_same",
                }
            ],
        },
    ]
    file_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    result = service.import_local_file("dataset2_question_json", str(file_path))
    detail = service.get_batch_detail(result.batch.id)

    assert detail.summary.failed == 0
    assert detail.summary.total_records == 2


def test_initialize_upload_batch_is_persisted_immediately() -> None:
    db = _build_session()
    _seed_import_context(db)
    service = ImportService(db)

    batch = service.initialize_upload_batch(
        "dataset2_question_json",
        file_name="题目合并 (10 files)",
    )

    reloaded = service.get_batch_detail(batch.id)
    assert reloaded.batch.id == batch.id
    assert reloaded.batch.import_status == "RUNNING"


def test_dataset1_import_creates_gold_label(tmp_path: Path) -> None:
    db = _build_session()
    _seed_import_context(db)
    service = ImportService(db)

    dataset2_path = tmp_path / "dataset2_base.json"
    _write_dataset2_json(
        dataset2_path,
        exercise_id="BASE_001",
        stem_text="$2022$ 的倒数是（ ） A. -2022 B. 2022 C. 1/2022 D. -1/2022",
    )
    service.import_local_file("dataset2_question_json", str(dataset2_path))

    dataset1_path = tmp_path / "dataset1.xlsx"
    _write_dataset1_xlsx(dataset1_path)
    result = service.import_local_file("dataset1_labeled", str(dataset1_path))
    detail = service.get_batch_detail(result.batch.id)

    assert detail.records[0].normalized_question_id is not None
    assert detail.records[0].parse_status in {"MATCHED_BY_CONTENT_HASH", "CREATED_NEW_QUESTION"}
    assert db.scalar(select(QuestionGoldLabel).limit(1)) is not None
    imported_question = db.scalar(select(Question).where(Question.id == detail.records[0].normalized_question_id))
    assert imported_question is not None
    assert imported_question.annotation_status == "COMPLETED"
    assert imported_question.annotation_count == imported_question.required_annotations


def test_dataset1_import_supports_senior_workbook_and_infers_grade(tmp_path: Path) -> None:
    db = _build_session()
    _seed_import_context(db)
    service = ImportService(db)

    dataset1_path = tmp_path / "dataset1_senior.xlsx"
    _write_dataset1_senior_xlsx(dataset1_path)
    result = service.import_local_file("dataset1_labeled", str(dataset1_path))
    detail = service.get_batch_detail(result.batch.id)

    assert detail.summary.failed == 0
    assert detail.summary.total_records == 2

    first_question = db.scalar(
        select(Question).join(QuestionExternalRef).where(QuestionExternalRef.external_question_id == "Q_4001")
    )
    second_question = db.scalar(
        select(Question).join(QuestionExternalRef).where(QuestionExternalRef.external_question_id == "Q_5001")
    )
    assert first_question is not None
    assert second_question is not None
    assert first_question.grade_id is not None
    assert second_question.grade_id is not None
    assert first_question.annotation_status == "COMPLETED"
    assert second_question.annotation_status == "COMPLETED"
    assert db.scalar(select(Grade.grade_name).where(Grade.id == first_question.grade_id)) == "高一"
    assert db.scalar(select(Grade.grade_name).where(Grade.id == second_question.grade_id)) == "高二"

    gold_label = db.scalar(
        select(QuestionGoldLabel).where(QuestionGoldLabel.question_id == second_question.id).limit(1)
    )
    assert gold_label is not None
    competency_rows = list(
        db.scalars(
            select(QuestionGoldCompetency).where(QuestionGoldCompetency.gold_label_id == gold_label.id)
        )
    )
    assert len(competency_rows) == 2
    assert sorted(item.level_value for item in competency_rows) == [1, 2]


def test_dataset3_import_creates_exam_response_domain_rows(tmp_path: Path) -> None:
    db = _build_session()
    _seed_import_context(db)
    service = ImportService(db)

    dataset3_path = tmp_path / "dataset3.xlsx"
    _write_dataset3_xlsx(dataset3_path)
    result = service.import_local_file("dataset3_exam_sheet", str(dataset3_path))
    detail = service.get_batch_detail(result.batch.id)

    assert detail.records[0].normalized_question_id is not None
    assert db.scalar(select(Exam).limit(1)) is not None
    assert db.scalar(select(StudentQuestionResponse).limit(1)) is not None
